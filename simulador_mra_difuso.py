"""
╔══════════════════════════════════════════════════════════════════════════════╗
║  SIMULADOR MRA — CONTROL DIFUSO MAMDANI   (Diseño Educativo GeoGebra)       ║
╚══════════════════════════════════════════════════════════════════════════════╝

PLANTA: Sistema Masa-Resorte-Amortiguador (vertical)
    m·ẍ(t) + b·ẋ(t) + k·x(t) = F(t)
    Integración Euler explícita,  dt = 0.02 s

CONTROLADOR: FIS Mamdani sin librerías externas
    Entradas: e(t) = xd − x   [−4,4] m
              Δe   = −ẋ(t)    [−6,6] m/s
    Salida  : F(t)            [−40,40] N
    5 conjuntos triangulares por variable, 25 reglas, defuzz. centroide

EXPORTACIÓN: Excel (.xlsx) con 3 hojas — Parámetros, Datos, Gráficas

RUBRICA CUCEA — Sistemas Dinámicos:
    ✓ Claridad conceptual  (modelado MRA + FIS Mamdani)
    ✓ Manejo del modelo    (Euler + 25 reglas + centroide)
    ✓ Lógica de simulación (Tkinter canvas, dot-plots, sliders)
    ✓ Gráfica de complemento: x(t) vs xd, e(t), F(t) en tiempo real
"""

import tkinter as tk
from tkinter import messagebox
import math
import os
import datetime
import threading
from collections import deque


# ══════════════════════════════════════════════════════════════════════════════
#  CONTROLADOR DIFUSO MAMDANI  (sin librerías externas)
# ══════════════════════════════════════════════════════════════════════════════

class FuzzyMRAController:
    """
    FIS Mamdani manual.
    Entradas : e (error), de (derivada del error)
    Salida   : F (fuerza de control)
    Conjuntos: 5 triangulares simétricos por variable
    AND      : mínimo  |  Defuzz: centroide singleton
    """

    def __init__(self):
        self.R_e  = 4.0    # radio universo error [m]
        self.R_de = 6.0    # radio universo Δe [m/s]
        self.R_F  = 40.0   # radio universo fuerza [N]

        # Tabla 5×5 — índice del conjunto de salida (0=NB…4=PB)
        # Centros de salida en orden INVERSO: 0→+40 N, 4→-40 N
        # (coherente con la física: e>0 → x<xd → F>0 para acercar masa)
        self.rules = [
            [4, 4, 3, 3, 2],   # e=NB
            [4, 3, 3, 2, 1],   # e=N
            [3, 3, 2, 1, 1],   # e=Z
            [3, 2, 1, 1, 0],   # e=P
            [2, 1, 1, 0, 0],   # e=PB
        ]

    @staticmethod
    def _tri(x, a, b, c):
        """Función de membresía triangular escalar."""
        lft = (x - a) / (b - a) if b != a else float(x >= b)
        rgt = (c - x) / (c - b) if c != b else float(x <= b)
        return max(0.0, min(lft, rgt))

    def _universe(self, R):
        """5 triángulos uniformes sobre [−R, R]."""
        h = R / 2.0
        return [(-R,-R,-h), (-R,-h,0.0), (-h,0.0,h), (0.0,h,R), (h,R,R)]

    def _fuzzify(self, val, R):
        val = max(-R, min(R, val))
        return [self._tri(val, a, b, c) for a, b, c in self._universe(R)]

    def infer(self, e, de):
        """Inferencia completa → fuerza F en Newtons."""
        mu_e  = self._fuzzify(e,  self.R_e)
        mu_de = self._fuzzify(de, self.R_de)

        # Centros singleton (orden inverso para coherencia física)
        raw_c = [b for (_, b, _) in self._universe(self.R_F)]  # [-40,-20,0,20,40]
        c_F   = list(reversed(raw_c))                           # [40, 20, 0,-20,-40]

        num, den = 0.0, 0.0
        for i in range(5):
            for j in range(5):
                alpha = min(mu_e[i], mu_de[j])
                if alpha > 1e-9:
                    num += alpha * c_F[self.rules[i][j]]
                    den += alpha

        if den < 1e-9:
            return 0.0
        return max(-self.R_F, min(self.R_F, num / den))


# ══════════════════════════════════════════════════════════════════════════════
#  PLANTA: SISTEMA MASA-RESORTE-AMORTIGUADOR
# ══════════════════════════════════════════════════════════════════════════════

class MRASystem:
    """
    m·ẍ + b·ẋ + k·x = F(t)
    Integración Euler explícita, dt = 0.02 s.
    x positivo = desplazamiento hacia abajo.
    """
    DT = 0.02

    def __init__(self, m=1.0, k=10.0, b=2.0, xd=1.0):
        self.m = m; self.k = k; self.b = b; self.xd = xd
        self.reset()

    def reset(self):
        self.x = 0.0; self.v = 0.0; self.t = 0.0
        self._peak_x  = 0.0    # para calcular sobrepaso
        self._t_settle = None  # tiempo de asentamiento

    def step(self, F):
        acc    = (F - self.b * self.v - self.k * self.x) / self.m
        self.v += acc * self.DT
        self.x += self.v * self.DT
        self.t += self.DT
        # Rastrear pico para sobrepaso
        if self.x > self._peak_x:
            self._peak_x = self.x
        return self.x, self.v, acc

    def omega_n(self):
        return math.sqrt(max(self.k, 1e-6) / max(self.m, 1e-6))

    def zeta(self):
        return self.b / (2.0 * math.sqrt(max(self.k * self.m, 1e-6)))

    def overshoot_pct(self):
        """Porcentaje de sobrepaso respecto al setpoint."""
        if abs(self.xd) < 1e-9:
            return 0.0
        return max(0.0, (self._peak_x - self.xd) / abs(self.xd) * 100.0)


# ══════════════════════════════════════════════════════════════════════════════
#  EXPORTACIÓN A EXCEL
# ══════════════════════════════════════════════════════════════════════════════

def exportar_excel(hist_t, hist_x, hist_v, hist_e, hist_F, m, k, b, xd, omega, zeta_val):
    """
    Genera un archivo .xlsx con 3 hojas:
      1. Parametros — m, k, b, xd, ωn, ζ, T, f
      2. Datos      — tabla t, x, e, v, F con formato de colores
      3. Graficas   — 3 gráficas de línea incrustadas (x+xd, e, F)
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.chart import LineChart, Reference
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None, "openpyxl no instalado. Ejecute: pip install openpyxl"

    wb = Workbook()

    # Estilos
    def fill(hex_c):
        return PatternFill("solid", fgColor=hex_c)
    thin  = Side(style='thin', color='BBBBBB')
    borde = Border(left=thin, right=thin, top=thin, bottom=thin)
    ctr   = Alignment(horizontal='center', vertical='center')

    # ── HOJA 1: Parámetros ────────────────────────────────────────────────────
    ws_p = wb.active
    ws_p.title = 'Parametros'
    T   = 2 * math.pi / omega if omega > 0 else 0
    f   = 1 / T if T > 0 else 0
    Mp  = max(0.0, (max(hist_x) - xd) / abs(xd) * 100) if abs(xd) > 1e-9 and hist_x else 0

    params = [
        ('Parámetro',          'Símbolo', 'Valor',              'Unidad'),
        ('Masa',               'm',       f'{m:.4f}',            'kg'),
        ('Constante resorte',  'k',       f'{k:.4f}',            'N/m'),
        ('Amortiguamiento',    'b',       f'{b:.4f}',            'Ns/m'),
        ('Setpoint',           'xd',      f'{xd:.4f}',           'm'),
        ('Frec. angular nat.', 'ωn',      f'{omega:.6f}',        'rad/s'),
        ('Coef. amort.',       'ζ',       f'{zeta_val:.6f}',     '—'),
        ('Período',            'T',       f'{T:.6f}',            's'),
        ('Frecuencia',         'f',       f'{f:.6f}',            'Hz'),
        ('Sobrepaso',          'Mp',      f'{Mp:.2f}',           '%'),
        ('Tipo de respuesta',  '—',
         'Subamortiguada' if zeta_val<1 else ('Crítica' if zeta_val==1 else 'Sobreamortiguada'), '—'),
        ('Puntos grabados',    'N',       str(len(hist_t)),      '—'),
        ('Generado',           '—',       datetime.datetime.now().strftime('%Y-%m-%d %H:%M'), '—'),
    ]

    hdrs_fill = ['1F4E79', '2E75B6', '9DC3E6', 'DDEBF7']
    for r, row in enumerate(params, 1):
        for c, val in enumerate(row, 1):
            cell = ws_p.cell(row=r, column=c, value=val)
            cell.border = borde; cell.alignment = ctr
            if r == 1:
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = fill(hdrs_fill[c-1])
            else:
                cell.fill = fill('EBF3FB' if r % 2 == 0 else 'FFFFFF')
    for c in range(1, 5):
        ws_p.column_dimensions[get_column_letter(c)].width = 26

    # ── HOJA 2: Datos ─────────────────────────────────────────────────────────
    ws_d = wb.create_sheet('Datos')

    cabeceras  = ['t (s)', 'x (m)', 'xd (m)', 'e=xd-x (m)', 'v (m/s)', 'F (N)']
    hdr_colors = ['333333', 'CC0000', '007700', '7B00CC',     '0033CC',  'FF6600']

    for c, (cab, col) in enumerate(zip(cabeceras, hdr_colors), 1):
        cell = ws_d.cell(row=1, column=c, value=cab)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = fill(col); cell.alignment = ctr; cell.border = borde
        ws_d.column_dimensions[get_column_letter(c)].width = 16

    for r, (t, x, e, v, F) in enumerate(
            zip(hist_t, hist_x, hist_e, hist_v, hist_F), 2):
        fila = [round(t,4), round(x,6), round(xd,4),
                round(e,6), round(v,6), round(F,4)]
        alt  = PatternFill("solid", fgColor='F8F8F8' if r % 2 == 0 else 'FFFFFF')
        for c, val in enumerate(fila, 1):
            cell = ws_d.cell(row=r, column=c, value=val)
            cell.border = borde; cell.alignment = ctr; cell.fill = alt

    ws_d.freeze_panes = 'A2'
    n_rows = len(hist_t)

    # ── HOJA 3: Gráficas ──────────────────────────────────────────────────────
    ws_g = wb.create_sheet('Graficas')

    def make_chart(title, cols, colors, row_off, height=10):
        chart = LineChart()
        chart.title = title; chart.style = 10
        chart.width = 24; chart.height = height
        chart.y_axis.title = title.split('(')[0].strip()
        chart.x_axis.title = 'Tiempo (s)'
        cats = Reference(ws_d, min_col=1, min_row=2, max_row=n_rows+1)
        for col_idx, color in zip(cols, colors):
            data = Reference(ws_d, min_col=col_idx, min_row=1, max_row=n_rows+1)
            chart.add_data(data, titles_from_data=True)
            s = chart.series[-1]
            s.graphicalProperties.line.solidFill = color
            s.graphicalProperties.line.width     = 18000
        chart.set_categories(cats)
        return chart

    # Gráfica 1: x(t) vs xd — columnas 2 y 3
    ch1 = make_chart('Posicion x(t) vs Setpoint xd  [m]',  [2, 3], ['CC0000','007700'], 1,  11)
    # Gráfica 2: error e(t)
    ch2 = make_chart('Error  e(t) = xd - x  [m]',          [4],    ['7B00CC'],          29, 10)
    # Gráfica 3: fuerza F(t)
    ch3 = make_chart('Fuerza de control  F(t)  [N]',        [6],    ['FF6600'],          48, 10)

    ws_g.add_chart(ch1, 'A1')
    ws_g.add_chart(ch2, 'A29')
    ws_g.add_chart(ch3, 'A48')

    # Guardar
    carpeta = os.path.join(os.path.expanduser('~'), 'Desktop',
                           'Sistema_Masa_Resorte')
    os.makedirs(carpeta, exist_ok=True)
    ts   = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    path = os.path.join(carpeta, f'MRA_Difuso_{ts}.xlsx')
    wb.save(path)
    return path, None


# ══════════════════════════════════════════════════════════════════════════════
#  APLICACIÓN GRÁFICA  (Tkinter — diseño GeoGebra)
# ══════════════════════════════════════════════════════════════════════════════

class SimuladorDifusoApp:
    """
    Interfaz visual estilo GeoGebra:

    ┌────────────┬───────────────────────┬──────────────────────────────────┐
    │ Sliders    │  Animación vertical   │  Dot-plots deslizantes:          │
    │  m,k,b,xd │  resorte helicoidal   │   x(t)  rojo  +  xd  verde       │
    │            │  + masa + xd line     │   e(t)  violeta                  │
    │            │  + flecha F           │   F(t)  naranja                  │
    └────────────┴───────────────────────┴──────────────────────────────────┘
        │  INICIO   PAUSA   REINICIO   │  EXPORTAR EXCEL  │  Métricas      │
    """

    # ── Dimensiones del canvas ───────────────────────────────────────────────
    W, H = 1100, 620

    # ── Zona de animación ────────────────────────────────────────────────────
    CX        = 285    # X central del sistema
    ANC_Y     = 80     # Y del punto de anclaje (techo)
    EQ_Y      = 270    # Y de la posición de equilibrio (x=0)
    PX_PER_M  = 70.0   # píxeles por metro

    # ── Zona de gráficas (derecha) ───────────────────────────────────────────
    GR_X0, GR_X1 = 450, 1060   # límites X
    GR_YC    = [115, 240, 365]  # Y central de cada gráfico (x, e, F)
    GR_AMP   = 50               # amplitud visual normalizada [px]
    WIN_T    = 10.0             # ventana temporal visible [s]

    # ── Sliders verticales ───────────────────────────────────────────────────
    # (x_track, y_top, y_bot, v_min, v_max, color_track, color_handle)
    SLIDER_DEFS = {
        "m" : ( 38, 110, 245,  0.5, 10.0, "#FF9800", "#E53935"),
        "k" : (105, 110, 245,  1.0, 50.0, "#9E9E9E", "#1E88E5"),
        "b" : ( 38, 290, 425,  0.1, 10.0, "#66BB6A", "#2E7D32"),
        "xd": (105, 290, 425, -2.5,  2.5, "#EF5350", "#880E4F"),
    }
    SLIDER_INIT = {"m": 1.0, "k": 10.0, "b": 2.0, "xd": 1.0}

    MAX_PTS = 200   # puntos máximos en historial

    # ── Constructor ──────────────────────────────────────────────────────────
    def __init__(self, root):
        self.root = root
        self.root.title("Sistema MRA — Control Difuso Mamdani  |  CUCEA")
        self.root.geometry(f"{self.W}x{self.H}")
        self.root.resizable(False, False)
        self.root.configure(bg="#1A1A1A")

        self.fis = FuzzyMRAController()
        self.sys = MRASystem()

        self.running    = False
        self.dragged    = None
        self.last_F     = 0.0
        self.sl_vals    = dict(self.SLIDER_INIT)
        self._exporting = False

        # Historial (deque O(1))
        self.h_t = deque(maxlen=self.MAX_PTS)
        self.h_x = deque(maxlen=self.MAX_PTS)
        self.h_v = deque(maxlen=self.MAX_PTS)
        self.h_e = deque(maxlen=self.MAX_PTS)   # error e(t)
        self.h_F = deque(maxlen=self.MAX_PTS)

        # Canvas principal
        self.canvas = tk.Canvas(
            self.root, width=self.W, height=self.H,
            bg="#FFF176", highlightthickness=0
        )
        self.canvas.pack(expand=True)

        self._build_static()
        self._build_dynamic()
        self._redraw()

        # Eventos
        self.canvas.bind("<Button-1>",        self._on_click)
        self.canvas.bind("<B1-Motion>",       self._on_drag)
        self.canvas.bind("<ButtonRelease-1>", self._on_release)
        self.canvas.bind("<Motion>",          self._on_hover)

        # Iniciar bucle
        self.root.after(20, self._tick)

    # ─────────────────────────────────────────────────────────────────────────
    #  ELEMENTOS ESTÁTICOS
    # ─────────────────────────────────────────────────────────────────────────
    def _build_static(self):
        c = self.canvas

        # Rejilla amarilla
        for x in range(0, self.W, 40):
            c.create_line(x, 0, x, self.H, fill="#E6D500", dash=(2, 6))
        for y in range(0, self.H, 40):
            c.create_line(0, y, self.W, y, fill="#E6D500", dash=(2, 6))

        # Título
        c.create_text(550, 22,
            text="Sistema Masa-Resorte-Amortiguador  —  Control Difuso Mamdani",
            font=("Arial", 14, "bold"), fill="#004D40")

        # ── Estructura metálica de soporte ────────────────────────────────
        sx = self.CX
        gray = "#666666"
        c.create_line(sx-95, 70, sx+95, 70,  fill=gray, width=7, capstyle="round")
        c.create_line(sx-95, 70, sx-95, 390, fill=gray, width=7, capstyle="round")
        c.create_line(sx+95, 70, sx+95, 390, fill=gray, width=7, capstyle="round")
        c.create_line(sx-95, 390,sx+95, 390, fill=gray, width=7, capstyle="round")
        c.create_line(sx-95, 390,sx-110, 430,fill=gray, width=7)
        c.create_line(sx+95, 390,sx+110, 430,fill=gray, width=7)
        # Articulaciones azules
        for jx, jy in [(sx,70),(sx-95,70),(sx+95,70),(sx-95,390),(sx+95,390)]:
            c.create_oval(jx-5,jy-5,jx+5,jy+5, fill="#4285F4",
                          outline="#212121", width=1.5)

        # Flecha Y (fija, morada)
        c.create_line(220, 245, 220, 160, fill="#7B1FA2",
                      width=2.5, arrow="last")
        c.create_text(208, 158, text="Y",
                      font=("Arial", 11, "bold"), fill="#7B1FA2")

        # ── Ejes de los 3 gráficos ────────────────────────────────────────
        x0, x1 = self.GR_X0, self.GR_X1
        # Eje vertical compartido
        c.create_line(x0, 45, x0, 430, fill="black", width=1.5)
        # Eje horizontal de cada gráfico + banda de fondo
        colores_banda = ["#FFEDED", "#F0EDFF", "#FFF0E0"]
        for yc, banda in zip(self.GR_YC, colores_banda):
            c.create_rectangle(x0, yc-self.GR_AMP-5, x1, yc+self.GR_AMP+5,
                               fill=banda, outline="")
            c.create_line(x0, yc, x1, yc, fill="#333333", width=1.5)

        # Etiquetas de señal
        lbl_data = [
            ("x / xd", "#CC0000"),
            ("e(t)",   "#7B00CC"),
            ("F(t)",   "#FF6600"),
        ]
        for yc, (lbl, col) in zip(self.GR_YC, lbl_data):
            c.create_text(x0-20, yc, text=lbl,
                          font=("Arial", 11, "bold"), fill=col, anchor="e")
            c.create_text(x1+18, yc, text="t",
                          font=("Arial", 11, "italic"), fill="black")

        # Separador entre zona de animación y gráficos
        c.create_line(435, 40, 435, 440, fill="#BBBBBB", width=1, dash=(4,4))

        # ── Panel de ecuaciones (parte inferior derecha) ───────────────────
        c.create_rectangle(450, 445, 1085, 540,
                           fill="white", outline="#AAAAAA", width=1)
        c.create_text(760, 453, anchor="n",
            text="Ecuaciones del sistema:",
            font=("Arial", 9, "bold"), fill="#004D40")
        ecuaciones = [
            ("m·ẍ + b·ẋ + k·x = F(t)",   "#333333", 460, 468),
            ("ωn = √(k/m)",                "#0033CC", 460, 482),
            ("ζ  = b / (2√(k·m))",         "#0033CC", 460, 496),
            ("e(t) = xd − x(t)",           "#7B00CC", 460, 510),
            ("F(t) = FIS_Mamdani(e, ė)",   "#CC0000", 460, 524),
            ("x(t)→xd  cuando  t→∞",       "#007700", 680, 468),
            ("Entradas FIS: e∈[−4,4]m",    "#555555", 680, 482),
            ("             de∈[−6,6]m/s",  "#555555", 680, 496),
            ("Salida FIS:   F∈[−40,40]N",  "#FF6600", 680, 510),
            ("25 reglas — Centroide",       "#555555", 680, 524),
        ]
        for txt, col, ex, ey in ecuaciones:
            c.create_text(ex, ey, text=txt, anchor="w",
                          font=("Courier", 9, "bold"), fill=col)

        # ── Tracks de los sliders ─────────────────────────────────────────
        sl_labels = {"m":"Masa","k":"Resorte","b":"Amortig.","xd":"Setpoint"}
        for name, (xt, yt, yb, _, _, col_t, _) in self.SLIDER_DEFS.items():
            c.create_line(xt, yt, xt, yb, fill=col_t, width=5)
            c.create_text(xt, yt-12, text=sl_labels[name],
                          font=("Arial", 8, "bold"), fill="#333333")

        # ── Botones de control (fondo blanco con borde coloreado) ─────────
        btns = [
            (115, 555, 202, 592, "INICIO",   "#CC0000", "btn_inicio"),
            (210, 555, 297, 592, "PAUSA",    "#0033CC", "btn_pausa"),
            (305, 555, 417, 592, "REINICIO", "#007700", "btn_reinicio"),
        ]
        for x1b,y1b,x2b,y2b,lbl,col,tag in btns:
            c.create_rectangle(x1b,y1b,x2b,y2b,
                               fill="white", outline=col, width=2.5,
                               tags=("btn", tag))
            c.create_text((x1b+x2b)//2, (y1b+y2b)//2,
                          text=lbl, font=("Arial", 10, "bold"),
                          fill=col, tags=("btn", tag))

        # Botón Excel (verde oscuro)
        c.create_rectangle(430, 555, 620, 592,
                           fill="#E8F5E9", outline="#2E7D32", width=2.5,
                           tags=("btn","btn_excel"))
        c.create_text(525, 573,
                      text="EXPORTAR EXCEL  ↓",
                      font=("Arial", 10, "bold"), fill="#1B5E20",
                      tags=("btn","btn_excel"))

        # Encabezado panel de sliders
        c.create_text(72, 95, text="PARÁMETROS",
                      font=("Arial", 9, "bold"), fill="#333333")

    # ─────────────────────────────────────────────────────────────────────────
    #  ELEMENTOS DINÁMICOS
    # ─────────────────────────────────────────────────────────────────────────
    def _build_dynamic(self):
        c = self.canvas

        # Etiquetas de valor de sliders
        self.id_lbl = {}
        lbl_cols = {"m":"#C62828","k":"#1565C0","b":"#2E7D32","xd":"#880E4F"}
        y_lbl    = {"m": 97,"k": 97,"b":278,"xd":278}
        for name in self.SLIDER_DEFS:
            self.id_lbl[name] = c.create_text(
                self.SLIDER_DEFS[name][0], y_lbl[name],
                text="", font=("Arial", 9, "bold"), fill=lbl_cols[name])

        # Handles de sliders
        self.id_hdl = {}
        for name, (*_, col_t, col_h) in self.SLIDER_DEFS.items():
            self.id_hdl[name] = c.create_oval(
                0,0,0,0, fill=col_h, outline="#212121", width=1.5)

        # ── Elementos de la animación ──────────────────────────────────────
        # Resorte helicoidal (smooth curve)
        self.id_spring = c.create_line(
            0,0,0,0, fill="#1565C0", width=3, smooth=True)

        # Masa esférica (roja con brillo)
        self.id_mass    = c.create_oval(0,0,0,0, fill="#D32F2F",
                                         outline="#1A1A1A", width=2)
        self.id_mass_hi = c.create_oval(0,0,0,0, fill="white", outline="")
        self.id_joint   = c.create_oval(0,0,0,0, fill="#4285F4",
                                         outline="#212121", width=1.5)

        # Línea de setpoint xd (verde punteada)
        self.id_xd_line = c.create_line(0,0,0,0, fill="#2E7D32",
                                         dash=(8,4), width=2)
        self.id_xd_lbl  = c.create_text(0,0, text="xd",
                                         font=("Arial",9,"bold"), fill="#2E7D32")

        # Flecha de fuerza F (naranja)
        self.id_F_arr = c.create_line(0,0,0,0, fill="#FF6600",
                                       width=3, arrow="last",
                                       arrowshape=(10,12,5))
        self.id_F_lbl = c.create_text(0,0, text="F",
                                       font=("Arial",11,"bold"), fill="#FF6600")

        # Línea de xd en gráfico x(t) (verde punteada horizontal)
        self.id_gxd = c.create_line(
            self.GR_X0, self.GR_YC[0],
            self.GR_X1, self.GR_YC[0],
            fill="#007700", dash=(6,3), width=1.5
        )

        # Texto de tiempo (dentro animación)
        self.id_time   = c.create_text(285, 420, text="t = 0.000 s",
                                        font=("Arial",12,"bold"), fill="#111111")
        # Métricas (ωn, ζ, e, Mp)
        self.id_metrics = c.create_text(285, 443, text="",
                                         font=("Arial", 9), fill="#333333")
        # Estado convergencia
        self.id_state   = c.create_text(285, 462, text="",
                                         font=("Arial", 10, "bold"), fill="#007700")
        # Estado del export
        self.id_export_st = c.create_text(525, 545, text="",
                                           font=("Arial", 8), fill="#1B5E20")

        # ── Dot-plots pre-creados ──────────────────────────────────────────
        # x(t): rojo  |  xd line: verde (constante, manejada aparte)
        # e(t): violeta  |  F(t): naranja
        dot_cfg = [("x","#CC0000"), ("e","#7B00CC"), ("F","#FF6600")]
        self.id_dots = {k: [] for k,_ in dot_cfg}
        for key, col in dot_cfg:
            for _ in range(self.MAX_PTS):
                self.id_dots[key].append(
                    c.create_oval(0,0,0,0, fill=col, outline="",
                                  state="hidden"))

    # ─────────────────────────────────────────────────────────────────────────
    #  REDIBUJADO
    # ─────────────────────────────────────────────────────────────────────────
    def _redraw(self):
        c   = self.canvas
        sys = self.sys
        sv  = self.sl_vals

        # ── Sliders ───────────────────────────────────────────────────────
        units = {"m":"kg","k":"N/m","b":"Ns/m","xd":"m"}
        for name, (xt, yt, yb, vmin, vmax, *_) in self.SLIDER_DEFS.items():
            val  = sv[name]
            frac = (val - vmin) / (vmax - vmin)
            yh   = yb - frac * (yb - yt)
            c.coords(self.id_hdl[name], xt-9, yh-9, xt+9, yh+9)
            c.itemconfig(self.id_lbl[name],
                         text=f"{name} = {val:.2f} {units[name]}")

        # ── Posición masa en pantalla ─────────────────────────────────────
        y_mass = self.EQ_Y + sys.x * self.PX_PER_M
        y_mass = max(120, min(385, y_mass))
        # Radio escala con la masa (visual)
        r_mass = 15 + sv["m"] / 10.0 * 8
        y_top  = y_mass - r_mass

        # ── Resorte helicoidal animado ────────────────────────────────────
        cx   = self.CX
        y_a  = self.ANC_Y + 8
        y_b  = y_top - 2
        pts  = [cx, self.ANC_Y, cx, y_a]
        if y_b > y_a + 4:
            n_coil, steps = 14, 130
            for i in range(steps + 1):
                frac  = i / steps
                y_p   = y_a + frac * (y_b - y_a)
                theta = frac * n_coil * 2 * math.pi
                x_p   = cx + 22 * math.sin(theta)
                pts.extend([x_p, y_p])
        pts.extend([cx, y_top])
        c.coords(self.id_spring, *pts)

        # ── Masa ──────────────────────────────────────────────────────────
        c.coords(self.id_mass,
                 cx-r_mass, y_mass-r_mass, cx+r_mass, y_mass+r_mass)
        hi = r_mass * 0.30
        c.coords(self.id_mass_hi,
                 cx-r_mass*0.40-hi, y_mass-r_mass*0.40-hi,
                 cx-r_mass*0.40+hi, y_mass-r_mass*0.40+hi)
        c.coords(self.id_joint, cx-5, y_top-5, cx+5, y_top+5)

        # ── Línea de setpoint xd ─────────────────────────────────────────
        y_xd = self.EQ_Y + sv["xd"] * self.PX_PER_M
        y_xd = max(120, min(385, y_xd))
        c.coords(self.id_xd_line, cx-90, y_xd, cx+90, y_xd)
        c.coords(self.id_xd_lbl,  cx+103, y_xd)

        # ── Flecha fuerza F ───────────────────────────────────────────────
        F = self.last_F
        if abs(F) > 0.5:
            pf = F * 1.2
            y_fend = max(50, min(500, y_mass + pf))
            c.coords(self.id_F_arr, cx, y_mass, cx, y_fend)
            c.coords(self.id_F_lbl, cx + 24, (y_mass + y_fend) / 2)
            c.itemconfig(self.id_F_arr, state="normal")
            c.itemconfig(self.id_F_lbl, state="normal")
        else:
            c.itemconfig(self.id_F_arr, state="hidden")
            c.itemconfig(self.id_F_lbl, state="hidden")

        # ── Métricas ──────────────────────────────────────────────────────
        wn = sys.omega_n(); z = sys.zeta()
        e  = sv["xd"] - sys.x
        Mp = sys.overshoot_pct()
        c.itemconfig(self.id_time, text=f"t = {sys.t:.2f} s")
        c.itemconfig(self.id_metrics,
            text=(f"ωn={wn:.3f} r/s   ζ={z:.3f}   "
                  f"e={e:+.3f}m   Mp={Mp:.1f}%"))
        if abs(e) < 0.05 and abs(sys.v) < 0.05 and sys.t > 0.1:
            c.itemconfig(self.id_state, text="✔ CONVERGIDO", fill="#007700")
        else:
            c.itemconfig(self.id_state, text="~ CONTROLANDO", fill="#CC0000")

        # ── Línea xd en gráfico x(t) ──────────────────────────────────────
        # Normalizar posición xd a píxeles del gráfico
        mx_sig = max((abs(v) for v in self.h_x), default=abs(sv["xd"])+0.01)
        mx_sig = max(mx_sig, abs(sv["xd"]) + 0.01)
        py_xd  = self.GR_YC[0] - (sv["xd"] / mx_sig) * self.GR_AMP
        c.coords(self.id_gxd, self.GR_X0, py_xd, self.GR_X1, py_xd)

        # ── Dot-plots deslizantes ─────────────────────────────────────────
        n     = len(self.h_t)
        t_now = sys.t
        x0, x1 = self.GR_X0, self.GR_X1
        yc_x, yc_e, yc_F = self.GR_YC

        mx_e = max((abs(v) for v in self.h_e), default=1.0) or 1.0
        mv_F = max((abs(v) for v in self.h_F), default=1.0) or 1.0

        for i in range(self.MAX_PTS):
            if i < n:
                dt_back = t_now - self.h_t[i]
                if 0.0 <= dt_back <= self.WIN_T:
                    px = x1 - (dt_back / self.WIN_T) * (x1 - x0)

                    # x(t) normalizado
                    py_x = yc_x - (self.h_x[i] / mx_sig) * self.GR_AMP
                    # e(t) normalizado
                    py_e = yc_e - (self.h_e[i] / mx_e) * self.GR_AMP
                    # F(t) normalizado
                    py_F = yc_F - (self.h_F[i] / mv_F) * self.GR_AMP

                    for key, py in [("x",py_x),("e",py_e),("F",py_F)]:
                        c.coords(self.id_dots[key][i], px-2,py-2, px+2,py+2)
                        c.itemconfig(self.id_dots[key][i], state="normal")
                    continue

            for key in ("x","e","F"):
                c.itemconfig(self.id_dots[key][i], state="hidden")

    # ─────────────────────────────────────────────────────────────────────────
    #  BUCLE DE SIMULACIÓN (20 ms → dt=0.02 s → tiempo real 1:1)
    # ─────────────────────────────────────────────────────────────────────────
    def _tick(self):
        if self.running:
            e  = self.sys.xd - self.sys.x
            de = -self.sys.v
            F  = self.fis.infer(e, de)
            self.last_F = F
            x, v, _ = self.sys.step(F)

            self.h_t.append(self.sys.t)
            self.h_x.append(x)
            self.h_v.append(v)
            self.h_e.append(e)
            self.h_F.append(F)

            self._redraw()

        self.root.after(20, self._tick)

    # ─────────────────────────────────────────────────────────────────────────
    #  PARÁMETROS Y RESET
    # ─────────────────────────────────────────────────────────────────────────
    def _apply_and_reset(self):
        sv = self.sl_vals
        self.sys.m = sv["m"]; self.sys.k = sv["k"]
        self.sys.b = sv["b"]; self.sys.xd = sv["xd"]
        self.sys.reset()
        self.h_t.clear(); self.h_x.clear()
        self.h_v.clear(); self.h_e.clear(); self.h_F.clear()
        self.last_F = 0.0

    # ─────────────────────────────────────────────────────────────────────────
    #  EXPORTAR EXCEL  (en hilo separado para no congelar la UI)
    # ─────────────────────────────────────────────────────────────────────────
    def _on_export(self):
        if self._exporting:
            return
        if len(self.h_t) < 5:
            self.canvas.itemconfig(self.id_export_st,
                                   text="Inicia la simulacion primero")
            return

        self._exporting = True
        self.canvas.itemconfig(self.id_export_st, text="Exportando...")

        # Captura de datos en el hilo principal antes de lanzar hilo
        data = (
            list(self.h_t), list(self.h_x), list(self.h_v),
            list(self.h_e), list(self.h_F),
            self.sys.m, self.sys.k, self.sys.b, self.sys.xd,
            self.sys.omega_n(), self.sys.zeta()
        )

        def worker():
            path, err = exportar_excel(*data)
            self._exporting = False
            if err:
                self.root.after(0, lambda: self.canvas.itemconfig(
                    self.id_export_st, text=f"Error: {err}"))
            else:
                nombre = os.path.basename(path)
                self.root.after(0, lambda: self.canvas.itemconfig(
                    self.id_export_st, text=f"Guardado: {nombre}"))
                try:
                    os.startfile(path)
                except Exception:
                    pass

        threading.Thread(target=worker, daemon=True).start()

    # ─────────────────────────────────────────────────────────────────────────
    #  EVENTOS DE RATÓN
    # ─────────────────────────────────────────────────────────────────────────
    def _slider_hit(self, mx, my):
        sv = self.sl_vals
        for name, (xt, yt, yb, vmin, vmax, *_) in self.SLIDER_DEFS.items():
            frac = (sv[name] - vmin) / (vmax - vmin)
            yh   = yb - frac * (yb - yt)
            if abs(mx - xt) <= 13 and abs(my - yh) <= 13:
                return name
        return None

    def _on_click(self, ev):
        x, y = ev.x, ev.y

        # Slider
        name = self._slider_hit(x, y)
        if name:
            self.dragged = name; return

        # INICIO
        if 115 <= x <= 202 and 555 <= y <= 592:
            self.running = True; return
        # PAUSA
        if 210 <= x <= 297 and 555 <= y <= 592:
            self.running = False; return
        # REINICIO
        if 305 <= x <= 417 and 555 <= y <= 592:
            self.running = False
            self._apply_and_reset()
            self._redraw(); return
        # EXCEL
        if 430 <= x <= 620 and 555 <= y <= 592:
            self._on_export(); return

    def _on_drag(self, ev):
        if not self.dragged: return
        name = self.dragged
        xt, yt, yb, vmin, vmax, *_ = self.SLIDER_DEFS[name]
        y_cl = max(yt, min(yb, ev.y))
        frac = (yb - y_cl) / (yb - yt)
        self.sl_vals[name] = vmin + frac * (vmax - vmin)
        self._apply_and_reset()
        self._redraw()

    def _on_release(self, ev):
        self.dragged = None

    def _on_hover(self, ev):
        x, y = ev.x, ev.y
        over = bool(self._slider_hit(x, y))
        if 115 <= x <= 620 and 555 <= y <= 592: over = True
        self.canvas.config(cursor="hand2" if over else "")


# ══════════════════════════════════════════════════════════════════════════════
#  PUNTO DE ENTRADA
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    root = tk.Tk()
    app  = SimuladorDifusoApp(root)
    root.mainloop()
