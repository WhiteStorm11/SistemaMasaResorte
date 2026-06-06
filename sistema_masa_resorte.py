"""
Sistema Masa-Resorte — Oscilador Armónico Simple
=================================================
Diseño visual inspirado en simuladores educativos:
  - Fondo amarillo, sistema vertical animado
  - Gráficas de posición, velocidad y aceleración en tiempo real
  - Sliders interactivos para m y k
  - Exportación completa a Excel (.xlsx) con datos + gráficas
"""

import numpy as np
import matplotlib
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import matplotlib.animation as animation
from matplotlib.widgets import Slider, Button
from matplotlib.gridspec import GridSpec
from collections import deque
import os, datetime

# ── Colores de la paleta del diseño ──────────────────────────
AMARILLO   = '#FFE800'
ROJO       = '#CC0000'
AZUL       = '#0033CC'
VERDE      = '#007700'
NARANJA    = '#FF6600'
GRIS_OSC   = '#333333'
GRIS_CLARO = '#888888'


# ══════════════════════════════════════════════════════════════
#  FÍSICA: OSCILADOR ARMÓNICO SIMPLE
# ══════════════════════════════════════════════════════════════

class MassSpringSystem:
    """
    Oscilador armónico simple vertical.

    x(t)  = x0 · sin(ω·t + φ)      posición respecto al equilibrio
    v(t)  = x0·ω · cos(ω·t + φ)    velocidad
    a(t)  = −x0·ω² · sin(ω·t + φ)  aceleración
    ω     = √(k / m)
    """

    def __init__(self, m=1.0, k=10.0, x0=0.15, phi=0.0, dt=0.02):
        self.m   = m        # masa (kg)
        self.k   = k        # constante del resorte (N/m)
        self.x0  = x0       # amplitud inicial (m)
        self.phi = phi      # fase inicial (rad)
        self.dt  = dt       # paso de tiempo (s)
        self.t   = 0.0

    @property
    def omega(self):
        """Frecuencia angular natural ω = √(k/m)."""
        return np.sqrt(max(self.k, 0.01) / max(self.m, 0.01))

    def state(self):
        """Retorna (x, v, a) en el instante actual."""
        wt = self.omega * self.t + self.phi
        x  =  self.x0            * np.sin(wt)
        v  =  self.x0 * self.omega * np.cos(wt)
        a  = -self.x0 * self.omega**2 * np.sin(wt)
        return x, v, a

    def step(self):
        """Avanza un paso dt y retorna (t, x, v, a)."""
        self.t += self.dt
        x, v, a = self.state()
        return self.t, x, v, a

    def reset(self):
        self.t = 0.0


# ══════════════════════════════════════════════════════════════
#  EXPORTACIÓN A EXCEL
# ══════════════════════════════════════════════════════════════

def exportar_excel(hist_t, hist_x, hist_v, hist_a, m, k, omega,
                   carpeta=None):
    """
    Genera un archivo .xlsx con:
      - Hoja 'Datos': tabla completa t, x, v, a
      - Hoja 'Parametros': m, k, ω, T, f
      - Hoja 'Graficas': gráficas de línea incrustadas
      - Formato con cabeceras coloreadas y anchos automáticos
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (Font, PatternFill, Alignment,
                                      Border, Side)
        from openpyxl.chart import LineChart, Reference
        from openpyxl.chart.series import SeriesLabel
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("ERROR: instale openpyxl  →  pip install openpyxl")
        return None

    wb = Workbook()

    # ── Estilos reutilizables ─────────────────────────────────
    def hdr_fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    thin = Side(style='thin', color='AAAAAA')
    borde = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center')

    # ══════════════════════════════════════════════════════════
    #  HOJA 1 — PARÁMETROS
    # ══════════════════════════════════════════════════════════
    ws_p = wb.active
    ws_p.title = 'Parametros'

    T  = 2 * np.pi / omega if omega > 0 else 0
    f  = 1 / T if T > 0 else 0

    params = [
        ('Parámetro', 'Símbolo', 'Valor', 'Unidad'),
        ('Masa',               'm',  f'{m:.4f}',         'kg'),
        ('Constante resorte',  'k',  f'{k:.4f}',         'N/m'),
        ('Frec. angular',      'ω',  f'{omega:.6f}',     'rad/s'),
        ('Período',            'T',  f'{T:.6f}',         's'),
        ('Frecuencia',         'f',  f'{f:.6f}',         'Hz'),
        ('Puntos grabados',    'N',  str(len(hist_t)),   '—'),
        ('Generado',           '—',  datetime.datetime.now().strftime('%Y-%m-%d %H:%M'), '—'),
    ]

    col_fills = ['1F4E79', '2E75B6', '9DC3E6', 'DDEBF7']  # azules
    for r, row in enumerate(params, 1):
        for c, val in enumerate(row, 1):
            cell = ws_p.cell(row=r, column=c, value=val)
            cell.border = borde
            cell.alignment = center
            if r == 1:
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = hdr_fill(col_fills[c - 1])
            else:
                cell.fill = hdr_fill('EBF3FB' if r % 2 == 0 else 'FFFFFF')

    for c in range(1, 5):
        ws_p.column_dimensions[get_column_letter(c)].width = 22

    # ══════════════════════════════════════════════════════════
    #  HOJA 2 — DATOS
    # ══════════════════════════════════════════════════════════
    ws_d = wb.create_sheet('Datos')

    cabeceras  = ['t (s)', 'x (m)', 'v (m/s)', 'a (m/s²)']
    hdr_colors = ['CC0000', '0033CC', '007700', 'FF6600']  # rojo, azul, verde, naranja

    for c, (cab, col) in enumerate(zip(cabeceras, hdr_colors), 1):
        cell = ws_d.cell(row=1, column=c, value=cab)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = hdr_fill(col)
        cell.alignment = center
        cell.border = borde
        ws_d.column_dimensions[get_column_letter(c)].width = 14

    for r, (t, x, v, a) in enumerate(zip(hist_t, hist_x, hist_v, hist_a), 2):
        datos = [round(t, 4), round(x, 6), round(v, 6), round(a, 6)]
        fill_row = PatternFill("solid", fgColor='F5F5F5' if r % 2 == 0 else 'FFFFFF')
        for c, val in enumerate(datos, 1):
            cell = ws_d.cell(row=r, column=c, value=val)
            cell.border = borde
            cell.alignment = center
            cell.fill = fill_row

    # Fijar encabezado al hacer scroll
    ws_d.freeze_panes = 'A2'

    # ══════════════════════════════════════════════════════════
    #  HOJA 3 — GRÁFICAS
    # ══════════════════════════════════════════════════════════
    ws_g = wb.create_sheet('Graficas')
    n_rows = len(hist_t)

    def make_chart(title, col_data, color, row_offset):
        """Crea un LineChart para la columna dada de la hoja Datos."""
        chart = LineChart()
        chart.title        = title
        chart.style        = 10
        chart.width        = 22
        chart.height       = 10
        chart.y_axis.title = title.split('(')[0].strip()
        chart.x_axis.title = 'Tiempo (s)'
        chart.x_axis.numFmt = '0.00'

        # Serie de datos
        data_ref = Reference(ws_d, min_col=col_data, min_row=1,
                             max_row=n_rows + 1)
        cats_ref = Reference(ws_d, min_col=1, min_row=2, max_row=n_rows + 1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)

        # Color de la línea
        s = chart.series[0]
        s.graphicalProperties.line.solidFill = color
        s.graphicalProperties.line.width     = 15000   # en EMUs (≈ 1.2 pt)
        s.smooth = False

        return chart

    configs = [
        ('Posicion  x(t)  [m]',      2, 'CC0000', 1),
        ('Velocidad  v(t)  [m/s]',   3, '0033CC', 25),
        ('Aceleracion  a(t)  [m/s²]',4, '007700', 49),
    ]
    for title, col, color, row_off in configs:
        ch = make_chart(title, col, color, row_off)
        ws_g.add_chart(ch, f'A{row_off}')

    # Guardar
    if carpeta is None:
        carpeta = os.path.join(os.path.expanduser('~'), 'Desktop',
                               'Sistema_Masa_Resorte')
    os.makedirs(carpeta, exist_ok=True)
    ts   = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    path = os.path.join(carpeta, f'MasaResorte_{ts}.xlsx')
    wb.save(path)
    return path


# ══════════════════════════════════════════════════════════════
#  APLICACIÓN GRÁFICA
# ══════════════════════════════════════════════════════════════

class SimuladorMasaResorte:
    """
    Interfaz gráfica completa del sistema masa-resorte.

    Layout (inspirado en la imagen de referencia):
    ┌──────────┬──────────────────┬──────────────────────┐
    │ Sliders  │  Animación 2D    │  Gráficas x, v, a    │
    │  m, k    │  (vertical)      │  (puntos de colores) │
    └──────────┴──────────────────┴──────────────────────┘
               │  Botones + Info  │
    """

    HIST    = 600    # puntos en historial
    DT      = 0.02   # paso de integración (s)
    X0      = 0.15   # amplitud inicial (m)

    def __init__(self):
        self.sys     = MassSpringSystem(m=1.0, k=10.0, x0=self.X0, dt=self.DT)
        self.running = False

        # Historial de señales
        self.hist_t = deque(maxlen=self.HIST)
        self.hist_x = deque(maxlen=self.HIST)
        self.hist_v = deque(maxlen=self.HIST)
        self.hist_a = deque(maxlen=self.HIST)

        self._build_figure()
        self._connect_widgets()

    # ── Figura principal ──────────────────────────────────────
    def _build_figure(self):
        self.fig = plt.figure(figsize=(14, 8))
        self.fig.patch.set_facecolor(AMARILLO)

        # Grid: 3 columnas (sliders | animación | gráficas)
        gs = GridSpec(
            1, 3, figure=self.fig,
            left=0.01, right=0.99, top=0.90, bottom=0.18,
            wspace=0.04, width_ratios=[0.18, 0.32, 0.50]
        )

        self.ax_sl   = self.fig.add_subplot(gs[0, 0])   # panel sliders
        self.ax_anim = self.fig.add_subplot(gs[0, 1])   # animación
        self.ax_grf  = self.fig.add_subplot(gs[0, 2])   # gráficas

        for ax in (self.ax_sl, self.ax_anim, self.ax_grf):
            ax.set_facecolor(AMARILLO)
            ax.set_xticks([]); ax.set_yticks([])
            for sp in ax.spines.values():
                sp.set_visible(False)

        # Título
        self.fig.text(0.5, 0.95, 'Sistema masa-resorte',
                      ha='center', va='center',
                      fontsize=18, fontweight='bold', color=VERDE,
                      fontfamily='sans-serif')

        self._build_sliders()
        self._build_animation()
        self._build_graphs()
        self._build_buttons()
        self._build_info_panel()

    # ── Sliders ───────────────────────────────────────────────
    def _build_sliders(self):
        """Sliders verticales de masa (rojo) y constante (verde-oliva)."""
        # Etiqueta m
        self.lbl_m = self.fig.text(
            0.025, 0.86, f'm = {self.sys.m:.1f} Kg',
            fontsize=11, fontweight='bold', color=ROJO
        )
        # Slider m (vertical)
        ax_m = self.fig.add_axes([0.03, 0.25, 0.02, 0.55],
                                   facecolor=AMARILLO)
        self.sl_m = Slider(ax_m, '', 0.1, 300.0, valinit=self.sys.m,
                           orientation='vertical', color=ROJO)
        self.sl_m.valtext.set_visible(False)
        self.sl_m.label.set_visible(False)

        # Etiqueta k
        self.lbl_k = self.fig.text(
            0.08, 0.60, f'k = {self.sys.k:.0f} N/m',
            fontsize=11, fontweight='bold', color='#556B00'
        )
        # Slider k (vertical)
        ax_k = self.fig.add_axes([0.10, 0.25, 0.02, 0.55],
                                   facecolor=AMARILLO)
        self.sl_k = Slider(ax_k, '', 1.0, 500.0, valinit=self.sys.k,
                           orientation='vertical', color='#8B9000')
        self.sl_k.valtext.set_visible(False)
        self.sl_k.label.set_visible(False)

    # ── Animación 2D vertical ─────────────────────────────────
    def _build_animation(self):
        ax = self.ax_anim
        ax.set_xlim(-1, 1)
        ax.set_ylim(-1.0, 1.2)

        # Marco/estructura gris
        estructura = mpatches.FancyBboxPatch(
            (-0.82, -0.85), 1.64, 1.95,
            boxstyle="square,pad=0.0",
            linewidth=3, edgecolor=GRIS_OSC, facecolor='none'
        )
        ax.add_patch(estructura)

        # Barra superior horizontal
        ax.plot([-0.82, 0.82], [1.1, 1.1], color=GRIS_OSC, lw=6, solid_capstyle='round')

        # Punto de anclaje (azul oscuro)
        self.anchor_dot, = ax.plot([0], [1.1], 'o', color='#00008B', ms=10, zorder=5)

        # Resorte (espiral simulada con línea zigzag)
        self.spring_line, = ax.plot([], [], color='#1E90FF', lw=2.5, zorder=4)

        # Masa (elipse roja/rosa)
        self.mass_ellipse = mpatches.Ellipse(
            (0, 0), 0.35, 0.20,
            facecolor='#FF6699', edgecolor=ROJO, lw=2.5, zorder=6
        )
        ax.add_patch(self.mass_ellipse)

        # Flecha Y (morada, hacia arriba, referencia)
        self.arr_Y = ax.annotate(
            '', xy=(0.12, 0.5), xytext=(0.12, 0.15),
            arrowprops=dict(arrowstyle='->', color='purple', lw=2)
        )
        ax.text(0.18, 0.52, 'Y', color='purple', fontsize=11, fontweight='bold')

        # Flecha F (verde, hacia abajo, peso)
        self.arr_F = ax.annotate(
            '', xy=(0, -0.55), xytext=(0, -0.25),
            arrowprops=dict(arrowstyle='->', color=VERDE, lw=2.5)
        )
        self.lbl_F = ax.text(0.06, -0.55, 'F', color=VERDE,
                              fontsize=12, fontweight='bold')

        # Línea de equilibrio (punteada)
        ax.axhline(0, color=GRIS_CLARO, lw=1, ls='--', alpha=0.6)

        # Etiqueta de tiempo
        self.lbl_tiempo = ax.text(
            -0.78, -0.80, 't = 0.000 s',
            fontsize=11, fontweight='bold', color=GRIS_OSC,
            bbox=dict(facecolor='white', edgecolor=GRIS_OSC, boxstyle='round,pad=0.3')
        )

        # Precomputar patrón del resorte
        self._n_coils = 12
        self._spring_t_norm = np.linspace(0, 1, self._n_coils * 2 + 2)
        self._spring_amp = 0.18
        self._spring_base_y = np.zeros(len(self._spring_t_norm))
        for i in range(1, len(self._spring_base_y) - 1):
            self._spring_base_y[i] = self._spring_amp if i % 2 == 1 else -self._spring_amp

    # ── Gráficas x(t), v(t), a(t) ────────────────────────────
    def _build_graphs(self):
        ax = self.ax_grf
        ax.set_xlim(0, 1)
        ax.set_ylim(-0.05, 1.05)

        # Líneas de separación de filas (estilo imagen)
        for y in [0.33, 0.66]:
            ax.axhline(y, color=GRIS_OSC, lw=1.2, alpha=0.4)

        # Ejes x, v, a virtuales (líneas centrales de cada franja)
        self._y_centers = [0.83, 0.50, 0.17]   # posición Y en coord normalizadas
        self._y_scales  = [0.13, 0.13, 0.13]   # amplitud normalizada

        labels = ['x', 'v', 'a']
        colors_lbl = [ROJO, AZUL, VERDE]
        for (y, lbl, col) in zip(self._y_centers, labels, colors_lbl):
            ax.axhline(y, color=GRIS_OSC, lw=0.7, ls='-', alpha=0.25)
            ax.text(-0.04, y + 0.03, lbl, transform=ax.transData,
                    fontsize=14, fontweight='bold', color=col)
            ax.text(1.01, y - 0.01, 't',
                    transform=ax.transData,
                    fontsize=11, color=GRIS_OSC, style='italic')

        # Scatter plots (puntos de colores como en la imagen)
        self.sc_x = ax.scatter([], [], s=6, c=ROJO,   zorder=5)
        self.sc_v = ax.scatter([], [], s=6, c=AZUL,   zorder=5)
        self.sc_a = ax.scatter([], [], s=6, c=VERDE,  zorder=5)

        # Ecuaciones debajo de las gráficas
        ecuaciones = (
            "Angulo de fase  φ = 0\n\n"
            "Posición      x(t) = x₀ · sen(ω · t)\n"
            "Velocidad     v(t) = v₀ · cos(ω · t)\n"
            "Aceleración   a(t) = −a₀ · sen(ω · t)"
        )
        self.fig.text(
            0.68, 0.125, ecuaciones,
            fontsize=9, color=GRIS_OSC, va='top',
            fontfamily='monospace',
            bbox=dict(facecolor='white', alpha=0.7,
                      edgecolor=GRIS_CLARO, boxstyle='round,pad=0.4')
        )

    # ── Botones ───────────────────────────────────────────────
    def _build_buttons(self):
        btn_style = dict(facecolor='white', hovercolor='#EEEEEE')
        pos_btns = [0.34, 0.44, 0.54]
        labels   = ['INICIO', 'PAUSA', 'REINICIO']
        colors   = [ROJO, AZUL, VERDE]
        self.btns = []
        for x_pos, lbl, col in zip(pos_btns, labels, colors):
            ax_b = self.fig.add_axes([x_pos, 0.04, 0.085, 0.07], **btn_style)
            b = Button(ax_b, lbl, color='white', hovercolor='#F0F0F0')
            b.label.set_color(col)
            b.label.set_fontsize(11)
            b.label.set_fontweight('bold')
            for sp in ax_b.spines.values():
                sp.set_edgecolor(col)
                sp.set_linewidth(2)
            self.btns.append(b)

        self.btn_inicio   = self.btns[0]
        self.btn_pausa    = self.btns[1]
        self.btn_reinicio = self.btns[2]

        # Botón exportar Excel
        ax_exp = self.fig.add_axes([0.68, 0.04, 0.12, 0.07], facecolor='white')
        self.btn_excel = Button(ax_exp, 'Exportar\nExcel',
                                color='white', hovercolor='#E8F5E9')
        self.btn_excel.label.set_color('#1D6B1D')
        self.btn_excel.label.set_fontsize(9)
        self.btn_excel.label.set_fontweight('bold')
        for sp in ax_exp.spines.values():
            sp.set_edgecolor('#1D6B1D')
            sp.set_linewidth(2)

        # Etiqueta de estado del export
        self.lbl_export = self.fig.text(
            0.83, 0.075, '', fontsize=8, color='#1D6B1D', va='center'
        )

    # ── Panel de información (ω, T, f) ───────────────────────
    def _build_info_panel(self):
        self.lbl_info = self.fig.text(
            0.155, 0.075, '',
            fontsize=9, color=GRIS_OSC, va='center',
            fontfamily='monospace',
            bbox=dict(facecolor='white', alpha=0.85,
                      edgecolor=GRIS_CLARO, boxstyle='round,pad=0.3')
        )
        self._refresh_info()

    def _refresh_info(self):
        w = self.sys.omega
        T = 2 * np.pi / w if w > 0 else 0
        self.lbl_info.set_text(
            f"ω = {w:.4f} rad/s\n"
            f"T = {T:.4f} s\n"
            f"f = {1/T:.4f} Hz" if T > 0 else "ω = 0"
        )

    # ── Conexión de eventos ───────────────────────────────────
    def _connect_widgets(self):
        self.sl_m.on_changed(self._on_m_change)
        self.sl_k.on_changed(self._on_k_change)
        self.btn_inicio.on_clicked(lambda _: self._set_running(True))
        self.btn_pausa.on_clicked(lambda _: self._set_running(False))
        self.btn_reinicio.on_clicked(self._on_reset)
        self.btn_excel.on_clicked(self._on_export)

    def _on_m_change(self, val):
        self.sys.m = val
        self.lbl_m.set_text(f'm = {val:.1f} Kg')
        self._refresh_info()

    def _on_k_change(self, val):
        self.sys.k = val
        self.lbl_k.set_text(f'k = {val:.0f} N/m')
        self._refresh_info()

    def _set_running(self, state: bool):
        self.running = state

    def _on_reset(self, _):
        self.running = False
        self.sys.reset()
        self.hist_t.clear(); self.hist_x.clear()
        self.hist_v.clear(); self.hist_a.clear()
        self.lbl_export.set_text('')
        self._update_animation(0.0)
        self.lbl_tiempo.set_text('t = 0.000 s')

    def _on_export(self, _):
        """Exporta los datos actuales a Excel."""
        if len(self.hist_t) < 2:
            self.lbl_export.set_text('Sin datos aun')
            return
        self.lbl_export.set_text('Exportando...')
        self.fig.canvas.draw_idle()

        path = exportar_excel(
            list(self.hist_t), list(self.hist_x),
            list(self.hist_v), list(self.hist_a),
            self.sys.m, self.sys.k, self.sys.omega
        )
        if path:
            nombre = os.path.basename(path)
            self.lbl_export.set_text(f'Guardado:\n{nombre}')
            # Abrir automáticamente en Windows
            try:
                os.startfile(path)
            except Exception:
                pass

    # ── Animación del resorte vertical ───────────────────────
    def _spring_zigzag(self, y_top: float, y_bottom: float):
        """Genera coordenadas x, y del resorte zigzag vertical."""
        ys = y_bottom + self._spring_t_norm * (y_top - y_bottom)
        xs = self._spring_base_y.copy()
        return xs, ys

    def _update_animation(self, x_pos: float):
        """Refresca la posición de la masa y el resorte."""
        # y_eq = 0 (equilibrio). Desplazamiento hacia abajo cuando x > 0
        y_masa = -x_pos * 2.0     # escala visual
        y_masa = float(np.clip(y_masa, -0.75, 0.85))

        # Resorte desde ancla (y=1.10) hasta borde superior de la masa
        y_top_spring = 1.10
        y_bot_spring = y_masa + 0.10  # borde superior de la elipse
        xs, ys = self._spring_zigzag(y_top_spring, y_bot_spring)
        self.spring_line.set_data(xs, ys)

        # Mover elipse (masa)
        self.mass_ellipse.set_center((0, y_masa))

        # Actualizar flecha F
        self.arr_F.xy     = (0, y_masa - 0.30)
        self.arr_F.xytext = (0, y_masa - 0.05)
        self.lbl_F.set_y(y_masa - 0.38)

    # ── Actualización de las gráficas de puntos ───────────────
    def _update_graphs(self):
        n = len(self.hist_t)
        if n < 2:
            return

        t_arr = np.array(self.hist_t)
        x_arr = np.array(self.hist_x)
        v_arr = np.array(self.hist_v)
        a_arr = np.array(self.hist_a)

        # Normalizar tiempo al eje [0,1] del axes
        t_min, t_max = t_arr[0], t_arr[-1]
        if t_max <= t_min:
            return
        t_n = (t_arr - t_min) / (t_max - t_min)

        # Normalizar cada señal a su franja vertical
        def norm_sig(sig, yc, yscale):
            mx = np.max(np.abs(sig)) if np.max(np.abs(sig)) > 1e-9 else 1.0
            return yc + (sig / mx) * yscale

        xn = norm_sig(x_arr, self._y_centers[0], self._y_scales[0])
        vn = norm_sig(v_arr, self._y_centers[1], self._y_scales[1])
        an = norm_sig(a_arr, self._y_centers[2], self._y_scales[2])

        self.sc_x.set_offsets(np.c_[t_n, xn])
        self.sc_v.set_offsets(np.c_[t_n, vn])
        self.sc_a.set_offsets(np.c_[t_n, an])

    # ── Callback de animación ─────────────────────────────────
    def _animate(self, _frame):
        if not self.running:
            return []

        t, x, v, a = self.sys.step()

        self.hist_t.append(t)
        self.hist_x.append(x)
        self.hist_v.append(v)
        self.hist_a.append(a)

        # Actualizar displays
        self.lbl_tiempo.set_text(f't = {t:.3f} s')
        self._update_animation(x)
        self._update_graphs()

        return []

    # ── Lanzador ─────────────────────────────────────────────
    def run(self):
        self.anim = animation.FuncAnimation(
            self.fig, self._animate,
            interval=20, blit=False, cache_frame_data=False
        )
        plt.show()


# ══════════════════════════════════════════════════════════════
#  PUNTO DE ENTRADA
# ══════════════════════════════════════════════════════════════

if __name__ == '__main__':
    app = SimuladorMasaResorte()
    app.run()
